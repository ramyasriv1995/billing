from datetime import datetime, timedelta
from tkinter import filedialog, messagebox

import customtkinter as ctk

from database import db
from .components import format_currency, page_header
from .theme import COLORS, FONT_BODY, FONT_HEADING, FONT_MONO, FONT_SMALL

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    Workbook = Font = PatternFill = None

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
except ImportError:
    colors = A4 = getSampleStyleSheet = None
    Paragraph = SimpleDocTemplate = Spacer = Table = TableStyle = None


class ReportsFrame(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent, fg_color=COLORS["bg"])
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)
        self.current_report = None
        self.report_kind = ctk.StringVar(value="Sales Report")
        self.period_var = ctk.StringVar(value="Monthly")
        self.start_var = ctk.StringVar()
        self.end_var = ctk.StringVar()
        self.opening_var = ctk.StringVar(value="0")
        self.expenses_var = ctk.StringVar(value="0")
        self._build()

    def refresh(self):
        return

    def _build(self):
        page_header(self, "Reports", "Generate professional sales, inventory and payment reports").grid(
            row=0, column=0, sticky="ew", padx=24, pady=(24, 12)
        )
        controls = ctk.CTkFrame(self, fg_color=COLORS["card"], corner_radius=12)
        controls.grid(row=1, column=0, sticky="ew", padx=24, pady=(0, 12))
        controls.grid_columnconfigure(6, weight=1)

        ctk.CTkLabel(controls, text="Report", font=FONT_BODY).grid(
            row=0, column=0, padx=(18, 6), pady=14
        )
        self.kind_combo = ctk.CTkComboBox(
            controls,
            values=["Sales Report", "Inventory Report", "Payment Report"],
            variable=self.report_kind,
            width=150,
            command=self._report_changed,
        )
        self.kind_combo.grid(row=0, column=1, padx=4, pady=14)

        ctk.CTkLabel(controls, text="Period", font=FONT_BODY).grid(
            row=0, column=2, padx=(12, 6), pady=14
        )
        self.period_combo = ctk.CTkComboBox(
            controls,
            values=["Daily", "Weekly", "Monthly", "Yearly", "Custom Date"],
            variable=self.period_var,
            width=125,
            command=lambda _: self._apply_period(),
        )
        self.period_combo.grid(row=0, column=3, padx=4, pady=14)

        self.start_entry = ctk.CTkEntry(controls, textvariable=self.start_var, width=110)
        self.start_entry.grid(row=0, column=4, padx=4, pady=14)
        self.end_entry = ctk.CTkEntry(controls, textvariable=self.end_var, width=110)
        self.end_entry.grid(row=0, column=5, padx=4, pady=14)

        ctk.CTkButton(
            controls, text="Generate", width=110, fg_color=COLORS["primary"],
            command=self._generate,
        ).grid(row=0, column=7, padx=4, pady=14)
        ctk.CTkButton(
            controls, text="Excel", width=80, command=self._export_excel
        ).grid(row=0, column=8, padx=4, pady=14)
        ctk.CTkButton(
            controls, text="PDF", width=80, command=self._export_pdf
        ).grid(row=0, column=9, padx=(4, 18), pady=14)

        self.cash_controls = ctk.CTkFrame(controls, fg_color="transparent")
        self.cash_controls.grid(row=1, column=0, columnspan=10, sticky="w", padx=18, pady=(0, 12))
        ctk.CTkLabel(self.cash_controls, text="Opening Balance", font=FONT_SMALL).pack(side="left")
        ctk.CTkEntry(self.cash_controls, textvariable=self.opening_var, width=100).pack(
            side="left", padx=(6, 16)
        )
        ctk.CTkLabel(self.cash_controls, text="Expenses", font=FONT_SMALL).pack(side="left")
        ctk.CTkEntry(self.cash_controls, textvariable=self.expenses_var, width=100).pack(
            side="left", padx=6
        )

        self.report_area = ctk.CTkScrollableFrame(
            self, fg_color=COLORS["card"], corner_radius=12
        )
        self.report_area.grid(row=2, column=0, sticky="nsew", padx=24, pady=(0, 24))
        self._apply_period()
        self._report_changed(self.report_kind.get())
        self._placeholder()

    def _report_changed(self, value):
        inventory = value == "Inventory Report"
        self.period_combo.configure(state="disabled" if inventory else "normal")
        self.start_entry.configure(state="disabled" if inventory else "normal")
        self.end_entry.configure(state="disabled" if inventory else "normal")
        if value == "Payment Report":
            self.cash_controls.grid()
        else:
            self.cash_controls.grid_remove()

    def _apply_period(self):
        today = datetime.now().date()
        period = self.period_var.get()
        if period == "Daily":
            start = end = today
        elif period == "Weekly":
            start, end = today - timedelta(days=today.weekday()), today
        elif period == "Monthly":
            start, end = today.replace(day=1), today
        elif period == "Yearly":
            start, end = today.replace(month=1, day=1), today
        else:
            if not self.start_var.get():
                self.start_var.set(today.strftime("%Y-%m-%d"))
            if not self.end_var.get():
                self.end_var.set(today.strftime("%Y-%m-%d"))
            return
        self.start_var.set(start.strftime("%Y-%m-%d"))
        self.end_var.set(end.strftime("%Y-%m-%d"))

    def _dates(self):
        for value in [self.start_var.get(), self.end_var.get()]:
            datetime.strptime(value, "%Y-%m-%d")
        return self.start_var.get(), self.end_var.get()

    def _generate(self):
        try:
            kind = self.report_kind.get()
            if kind == "Inventory Report":
                report = db.get_inventory_report()
            else:
                start, end = self._dates()
                if kind == "Sales Report":
                    report = db.get_sales_report(start, end)
                else:
                    report = db.get_payment_report(
                        start, end, float(self.opening_var.get() or 0),
                        float(self.expenses_var.get() or 0),
                    )
            self.current_report = report
            self._render(report)
        except ValueError:
            messagebox.showerror("Invalid values", "Use YYYY-MM-DD dates and valid numeric balances.")

    def _placeholder(self):
        ctk.CTkLabel(
            self.report_area, text="Choose a report and click Generate.",
            font=FONT_BODY, text_color=COLORS["text_muted"],
        ).pack(pady=50)

    def _clear(self):
        for widget in self.report_area.winfo_children():
            widget.destroy()

    def _render(self, report):
        self._clear()
        ctk.CTkLabel(
            self.report_area,
            text=f"{report['report_type']} — {report['period_label']}",
            font=("Helvetica", 20, "bold"),
        ).pack(anchor="w", padx=24, pady=(20, 4))
        if report["report_type"] == "Sales Report":
            self._cards([
                ("Revenue", report["revenue"]), ("Profit", report["profit"]),
                ("Tax Collected", report["tax_collected"]), ("Invoices", report["sales_count"]),
            ])
            self._table("Product Wise Sales", ["Product", "Qty", "Revenue", "Profit"], [
                [r["product_name"], f"{r['quantity']:g}", format_currency(r["revenue"]),
                 format_currency(r["profit"])] for r in report["product_sales"]
            ])
        elif report["report_type"] == "Inventory Report":
            self._cards([
                ("Current Stock", report["total_units"]),
                ("Low Stock", len(report["low_stock"])),
                ("Inventory Value", report["inventory_value"]),
                ("Out of Stock", len(report["out_of_stock"])),
            ])
            self._table("Current Stock", ["Product", "Stock", "Unit", "Value"], [
                [p["name"], f"{p['stock_quantity']:g}", p.get("unit", ""),
                 format_currency(p["stock_quantity"] * p["cost_price"])]
                for p in report["current_stock"]
            ])
            self._table("Low Stock Products", ["Product", "Stock", "Minimum"], [
                [p["name"], f"{p['stock_quantity']:g}", f"{p['min_stock_level']:g}"]
                for p in report["low_stock"]
            ])
            self._table("Out of Stock Products", ["Product", "Unit"], [
                [p["name"], p.get("unit", "")] for p in report["out_of_stock"]
            ])
        else:
            collections = report["collections"]
            self._cards([
                ("Cash", collections.get("Cash", 0)), ("UPI", collections.get("UPI", 0)),
                ("Card", collections.get("Card", 0)),
                ("Bank Transfer", collections.get("Bank Transfer", 0)),
            ])
            self._table("Daily Cash Summary", ["Opening", "Sales", "Expenses", "Closing"], [[
                format_currency(report["opening_balance"]),
                format_currency(report["sales_collection"]),
                format_currency(report["expenses"]),
                format_currency(report["closing_balance"]),
            ]])

    def _cards(self, values):
        frame = ctk.CTkFrame(self.report_area, fg_color=COLORS["bg"], corner_radius=10)
        frame.pack(fill="x", padx=24, pady=14)
        for index, (label, value) in enumerate(values):
            frame.grid_columnconfigure(index, weight=1)
            card = ctk.CTkFrame(frame, fg_color="transparent")
            card.grid(row=0, column=index, padx=12, pady=14)
            ctk.CTkLabel(card, text=label, font=FONT_SMALL, text_color=COLORS["text_muted"]).pack()
            display = format_currency(value) if isinstance(value, float) else str(value)
            ctk.CTkLabel(card, text=display, font=FONT_HEADING).pack(pady=(4, 0))

    def _table(self, title, headers, rows):
        ctk.CTkLabel(self.report_area, text=title, font=FONT_HEADING).pack(
            anchor="w", padx=24, pady=(10, 6)
        )
        table = ctk.CTkFrame(self.report_area, fg_color=COLORS["bg"], corner_radius=8)
        table.pack(fill="x", padx=24, pady=(0, 14))
        header = ctk.CTkFrame(table, fg_color=COLORS["sidebar"])
        header.pack(fill="x", padx=8, pady=(8, 3))
        width = max(130, int(700 / max(len(headers), 1)))
        for value in headers:
            ctk.CTkLabel(header, text=value, text_color="white", font=FONT_BODY,
                         width=width, anchor="w").pack(side="left", padx=8, pady=7)
        for values in rows:
            row = ctk.CTkFrame(table, fg_color="transparent")
            row.pack(fill="x", padx=8, pady=2)
            for value in values:
                ctk.CTkLabel(row, text=str(value), font=FONT_MONO, width=width,
                             anchor="w").pack(side="left", padx=8, pady=5)
        if not rows:
            ctk.CTkLabel(table, text="No records found", text_color=COLORS["text_muted"]).pack(pady=14)

    def _export_rows(self):
        report = self.current_report
        if report["report_type"] == "Sales Report":
            return ["Product", "Quantity", "Revenue", "Profit"], [
                [r["product_name"], r["quantity"], r["revenue"], r["profit"]]
                for r in report["product_sales"]
            ]
        if report["report_type"] == "Inventory Report":
            return ["Product", "Stock", "Unit", "Inventory Value"], [
                [p["name"], p["stock_quantity"], p.get("unit", ""),
                 p["stock_quantity"] * p["cost_price"]] for p in report["current_stock"]
            ]
        return ["Method", "Collection"], list(report["collections"].items())

    def _export_excel(self):
        if not self.current_report or Workbook is None:
            messagebox.showwarning("Generate report", "Generate a report before exporting.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
            initialfile=self.current_report["report_type"].lower().replace(" ", "_") + ".xlsx",
        )
        if not path:
            return
        headers, rows = self._export_rows()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Report"
        sheet.append([self.current_report["report_type"]])
        sheet.append([self.current_report["period_label"]])
        sheet.append([])
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        for cell in sheet[4]:
            cell.fill = PatternFill("solid", fgColor="2563EB")
            cell.font = Font(color="FFFFFF", bold=True)
        workbook.save(path)
        messagebox.showinfo("Exported", f"Excel saved to:\n{path}")

    def _export_pdf(self):
        if not self.current_report or SimpleDocTemplate is None:
            messagebox.showwarning("Generate report", "Generate a report before exporting.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".pdf", filetypes=[("PDF", "*.pdf")],
            initialfile=self.current_report["report_type"].lower().replace(" ", "_") + ".pdf",
        )
        if not path:
            return
        headers, rows = self._export_rows()
        styles = getSampleStyleSheet()
        story = [
            Paragraph(self.current_report["report_type"], styles["Title"]),
            Paragraph(self.current_report["period_label"], styles["BodyText"]),
            Spacer(1, 16),
        ]
        table = Table([headers] + rows, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        story.append(table)
        SimpleDocTemplate(path, pagesize=A4).build(story)
        messagebox.showinfo("Exported", f"PDF saved to:\n{path}")
