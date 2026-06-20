import csv
import os
import subprocess
import tempfile
from datetime import datetime
from tkinter import filedialog, messagebox

import customtkinter as ctk

from database import db
from .components import format_currency, page_header, styled_treeview
from .theme import COLORS, FONT_BODY, FONT_HEADING, FONT_MONO, FONT_SMALL

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:
    Workbook = None
    Alignment = Border = Font = PatternFill = Side = None

try:
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
except ImportError:
    rl_colors = None
    A4 = letter = None
    getSampleStyleSheet = None
    mm = None
    Paragraph = SimpleDocTemplate = Spacer = Table = TableStyle = None


class InvoiceDetailsFrame(ctk.CTkFrame):
    def __init__(self, parent, highlight_sale_id=None):
        super().__init__(parent, fg_color=COLORS["bg"])
        self.highlight_sale_id = highlight_sale_id
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)
        self._build()

    def refresh(self, highlight_sale_id=None):
        if highlight_sale_id:
            self.highlight_sale_id = highlight_sale_id
        for widget in self.winfo_children():
            widget.destroy()
        self._build()

    def _build(self):
        header_row = ctk.CTkFrame(self, fg_color="transparent")
        header_row.grid(row=0, column=0, sticky="ew", padx=24, pady=(24, 8))
        header_row.grid_columnconfigure(0, weight=1)

        page_header(header_row, "Invoice Details", "View saved invoices and sales history").grid(
            row=0, column=0, sticky="w"
        )
        ctk.CTkButton(
            header_row,
            text="View Invoice",
            width=140,
            fg_color=COLORS["primary"],
            command=self._view_invoice,
        ).grid(row=0, column=1, sticky="e", padx=4)

        search_row = ctk.CTkFrame(self, fg_color="transparent")
        search_row.grid(row=1, column=0, sticky="ew", padx=24, pady=(0, 8))
        self.search_var = ctk.StringVar()
        self.search_var.trace_add("write", lambda *_: self._load_sales())
        ctk.CTkEntry(
            search_row,
            placeholder_text="Search by invoice or customer...",
            width=320,
            textvariable=self.search_var,
        ).pack(side="left")

        cols = ("id", "invoice", "customer", "total", "payment", "date")
        headings = {
            "id": "ID",
            "invoice": "Invoice #",
            "customer": "Customer",
            "total": "Total",
            "payment": "Payment",
            "date": "Date",
        }
        widths = {
            "id": 50,
            "invoice": 140,
            "customer": 180,
            "total": 110,
            "payment": 110,
            "date": 150,
        }
        self.tree, tree_container = styled_treeview(self, cols, headings, widths)
        tree_container.grid(row=2, column=0, sticky="nsew", padx=24, pady=(0, 24))

        self._load_sales()

        if self.highlight_sale_id:
            for item in self.tree.get_children():
                if str(self.tree.item(item)["values"][0]) == str(self.highlight_sale_id):
                    self.tree.selection_set(item)
                    self.tree.see(item)
                    self._view_invoice()
                    break
            self.highlight_sale_id = None

    def _load_sales(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        search = self.search_var.get() if hasattr(self, "search_var") else ""
        for s in db.get_sales(search):
            self.tree.insert(
                "",
                "end",
                values=(
                    s["id"],
                    s["invoice_number"],
                    s.get("customer_name") or "Walk-in",
                    format_currency(s["total"]),
                    s["payment_method"],
                    s["created_at"][:16],
                ),
            )

    def _get_selected_id(self):
        sel = self.tree.selection()
        if not sel:
            return None
        return str(self.tree.item(sel[0])["values"][0])

    def _view_invoice(self):
        sale_id = self._get_selected_id()
        if not sale_id:
            return
        sale = db.get_sale(sale_id)
        if not sale:
            return

        dialog = ctk.CTkToplevel(self)
        dialog.title(f"Invoice {sale['invoice_number']}")
        dialog.geometry("560x660")
        dialog.resizable(False, False)
        dialog.transient(self.winfo_toplevel())
        dialog.grab_set()

        content = ctk.CTkScrollableFrame(dialog, fg_color=COLORS["card"], width=520, height=620)
        content.pack(fill="both", expand=True, padx=16, pady=16)

        ctk.CTkLabel(content, text="INVOICE", font=("Helvetica", 24, "bold")).pack(
            anchor="w", pady=(8, 4)
        )
        ctk.CTkLabel(
            content,
            text=sale["invoice_number"],
            font=FONT_HEADING,
            text_color=COLORS["primary"],
        ).pack(anchor="w")
        ctk.CTkLabel(
            content,
            text=f"Date: {sale['created_at']}",
            font=FONT_SMALL,
            text_color=COLORS["text_muted"],
        ).pack(anchor="w", pady=(0, 16))

        ctk.CTkLabel(content, text="Bill To:", font=FONT_HEADING).pack(anchor="w")
        customer = sale.get("customer_name") or "Walk-in Customer"
        ctk.CTkLabel(content, text=customer, font=FONT_BODY).pack(anchor="w")
        if sale.get("customer_phone"):
            ctk.CTkLabel(content, text=sale["customer_phone"], font=FONT_SMALL).pack(anchor="w")
        if sale.get("customer_email"):
            ctk.CTkLabel(content, text=sale["customer_email"], font=FONT_SMALL).pack(anchor="w")
        if sale.get("customer_address"):
            ctk.CTkLabel(content, text=sale["customer_address"], font=FONT_SMALL).pack(anchor="w")
        if sale.get("customer_gst"):
            ctk.CTkLabel(content, text=f"GST: {sale['customer_gst']}", font=FONT_SMALL).pack(anchor="w")

        ctk.CTkFrame(content, height=2, fg_color=COLORS["border"]).pack(fill="x", pady=16)

        header = ctk.CTkFrame(content, fg_color=COLORS["bg"], corner_radius=6)
        header.pack(fill="x", pady=(0, 4))
        for text, w in [
            ("#", 30), ("Item", 180), ("Qty", 45), ("Unit", 55), ("Price", 75), ("Total", 80)
        ]:
            ctk.CTkLabel(header, text=text, font=FONT_HEADING, width=w, anchor="w").pack(
                side="left", padx=8, pady=8
            )

        for index, item in enumerate(sale["items"], start=1):
            row = ctk.CTkFrame(content, fg_color="transparent")
            row.pack(fill="x", pady=2)
            for text, w in [
                (str(index), 30),
                (item["product_name"], 180),
                (f"{item['quantity']:g}", 45),
                (item.get("unit", "pcs"), 55),
                (format_currency(item["unit_price"]), 75),
                (format_currency(item.get("amount", item["subtotal"])), 80),
            ]:
                ctk.CTkLabel(row, text=text, font=FONT_MONO, width=w, anchor="w").pack(
                    side="left", padx=8
                )

        ctk.CTkFrame(content, height=2, fg_color=COLORS["border"]).pack(fill="x", pady=16)

        totals_frame = ctk.CTkFrame(content, fg_color="transparent")
        totals_frame.pack(fill="x")

        for label, value in [
            ("Subtotal:", format_currency(sale["subtotal"])),
            (f"Tax ({sale['tax_rate']}%):", format_currency(sale["tax_amount"])),
            ("Discount:", format_currency(sale["discount"])),
            ("Total:", format_currency(sale["total"])),
        ]:
            row = ctk.CTkFrame(totals_frame, fg_color="transparent")
            row.pack(fill="x", pady=2)
            ctk.CTkLabel(row, text=label, font=FONT_BODY, anchor="e", width=200).pack(side="left")
            font = ("Helvetica", 16, "bold") if label == "Total:" else FONT_BODY
            color = COLORS["primary"] if label == "Total:" else COLORS["text"]
            ctk.CTkLabel(row, text=value, font=font, text_color=color).pack(side="right")

        ctk.CTkLabel(
            content,
            text=f"Payment Method: {sale['payment_method']}",
            font=FONT_SMALL,
            text_color=COLORS["text_muted"],
        ).pack(anchor="w", pady=(16, 8))

        ctk.CTkButton(dialog, text="Close", width=100, command=dialog.destroy).pack(pady=12)


class BillingFrame(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent, fg_color=COLORS["bg"])
        self.cart = []
        self._customers = {}
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self._build()

    def refresh(self):
        self.cart.clear()
        for widget in self.winfo_children():
            widget.destroy()
        self._build()

    def _build(self):
        page_header(self, "Billing", "Create, edit, export and print customer invoices").grid(
            row=0, column=0, sticky="ew", padx=24, pady=(24, 12)
        )

        content = ctk.CTkScrollableFrame(
            self,
            fg_color=COLORS["card"],
            corner_radius=12,
            scrollbar_button_color="#F2AA2E",
            scrollbar_button_hover_color="#D99116",
        )
        content.grid(row=1, column=0, sticky="nsew", padx=24, pady=(0, 24))
        content.grid_columnconfigure(0, weight=1)

        invoice_banner = ctk.CTkFrame(
            content, fg_color="#292A33", corner_radius=10, height=124
        )
        invoice_banner.grid(row=0, column=0, sticky="ew", padx=16, pady=(16, 14))
        invoice_banner.grid_columnconfigure(0, weight=1)
        invoice_banner.grid_propagate(False)

        brand_block = ctk.CTkFrame(invoice_banner, fg_color="transparent")
        brand_block.grid(row=0, column=0, sticky="w", padx=24, pady=20)
        ctk.CTkLabel(
            brand_block, text="BILLING & INVENTORY",
            font=("Helvetica", 15, "bold"), text_color="white",
        ).pack(anchor="w")
        ctk.CTkLabel(
            brand_block, text="Professional customer invoice",
            font=FONT_SMALL, text_color="#D1D5DB",
        ).pack(anchor="w", pady=(5, 0))

        title_block = ctk.CTkFrame(invoice_banner, fg_color="transparent")
        title_block.grid(row=0, column=1, sticky="e", padx=28, pady=16)
        ctk.CTkLabel(
            title_block, text="CUSTOMER BILL",
            font=("Helvetica", 15, "italic"), text_color="white",
        ).pack(anchor="e")
        ctk.CTkLabel(
            title_block, text="INVOICE",
            font=("Helvetica", 34, "bold"), text_color="#F2AA2E",
        ).pack(anchor="e")

        top = ctk.CTkFrame(content, fg_color="transparent")
        top.grid(row=1, column=0, sticky="ew", padx=16, pady=(0, 14))
        top.grid_columnconfigure(0, weight=1)
        top.grid_columnconfigure(1, weight=1)

        left_panel = ctk.CTkFrame(
            top, fg_color="#FFFFFF", corner_radius=8,
            border_width=1, border_color=COLORS["border"],
        )
        left_panel.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        right_panel = ctk.CTkFrame(
            top, fg_color="#FFFFFF", corner_radius=8,
            border_width=1, border_color=COLORS["border"],
        )
        right_panel.grid(row=0, column=1, sticky="nsew", padx=(8, 0))

        # Customer section
        ctk.CTkLabel(
            left_panel, text="BILL TO", font=FONT_HEADING,
            fg_color="#F2AA2E", text_color="#292A33",
            corner_radius=4, padx=12, pady=4,
        ).pack(anchor="w", padx=16, pady=(14, 8))

        customers = db.get_customers()
        self._customers = {c["name"]: c["id"] for c in customers}
        cust_values = ["Walk-in Customer"] + [c["name"] for c in customers]
        self.customer_combo = ctk.CTkComboBox(
            left_panel,
            values=cust_values,
            width=300,
            command=self._apply_customer,
        )
        self.customer_combo.set("Walk-in Customer")
        self.customer_combo.pack(fill="x", padx=16, pady=(0, 10))

        self.customer_name = ctk.StringVar(value="")
        self.customer_phone = ctk.StringVar(value="")
        self.customer_email = ctk.StringVar(value="")
        self.customer_gst = ctk.StringVar(value="")
        self.customer_address = ctk.StringVar(value="")
        self.save_customer_var = ctk.BooleanVar(value=False)

        customer_grid = ctk.CTkFrame(left_panel, fg_color="transparent")
        customer_grid.pack(fill="x", padx=16)
        customer_grid.grid_columnconfigure(0, weight=1)
        customer_grid.grid_columnconfigure(1, weight=1)
        for col, label, variable in [
            (0, "Customer Name", self.customer_name),
            (1, "Phone Number", self.customer_phone),
        ]:
            ctk.CTkLabel(customer_grid, text=label, font=FONT_SMALL).grid(
                row=0, column=col, sticky="w", padx=(0, 8)
            )
            ctk.CTkEntry(customer_grid, textvariable=variable, width=180).grid(
                row=1, column=col, sticky="ew", padx=(0, 8), pady=(0, 6)
            )
        for col, label, variable in [
            (0, "Email Address", self.customer_email),
            (1, "GST Number", self.customer_gst),
        ]:
            ctk.CTkLabel(customer_grid, text=label, font=FONT_SMALL).grid(
                row=2, column=col, sticky="w", padx=(0, 8)
            )
            ctk.CTkEntry(customer_grid, textvariable=variable, width=180).grid(
                row=3, column=col, sticky="ew", padx=(0, 8), pady=(0, 6)
            )
        ctk.CTkLabel(left_panel, text="Address", font=FONT_SMALL).pack(
            anchor="w", padx=16
        )
        self.customer_address_box = ctk.CTkTextbox(left_panel, height=58, wrap="word")
        self.customer_address_box.pack(fill="x", padx=16, pady=(0, 4))
        ctk.CTkCheckBox(
            left_panel,
            text="Save these customer details",
            variable=self.save_customer_var,
            font=FONT_SMALL,
        ).pack(anchor="w", padx=16, pady=(0, 14))

        # Bill information section
        ctk.CTkLabel(
            right_panel, text="INVOICE DETAILS", font=FONT_HEADING,
            fg_color="#F2AA2E", text_color="#292A33",
            corner_radius=4, padx=12, pady=4,
        ).pack(anchor="w", padx=16, pady=(14, 8))
        self.invoice_number = ctk.StringVar(value=db.generate_invoice_number())
        self.invoice_date = ctk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        self.payment_var = ctk.StringVar(value="Cash")
        self.tax_rate_var = ctk.StringVar(value="0")
        self.discount_var = ctk.StringVar(value="0")
        self.cash_received_var = ctk.StringVar(value="")

        info_grid = ctk.CTkFrame(right_panel, fg_color="transparent")
        info_grid.pack(fill="x", padx=16, pady=(0, 14))
        info_grid.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(info_grid, text="Invoice #", font=FONT_SMALL).grid(row=0, column=0, sticky="w", pady=4)
        ctk.CTkEntry(info_grid, textvariable=self.invoice_number, width=220).grid(row=0, column=1, sticky="w", pady=4)

        ctk.CTkLabel(info_grid, text="Invoice Date", font=FONT_SMALL).grid(row=1, column=0, sticky="w", pady=4)
        ctk.CTkEntry(info_grid, textvariable=self.invoice_date, width=220).grid(row=1, column=1, sticky="w", pady=4)

        ctk.CTkLabel(info_grid, text="Payment Mode", font=FONT_SMALL).grid(row=2, column=0, sticky="w", pady=4)
        ctk.CTkComboBox(
            info_grid,
            values=["Cash", "UPI", "Card", "Bank Transfer"],
            variable=self.payment_var,
            width=220,
        ).grid(row=2, column=1, sticky="w", pady=4)

        # Product area
        body = ctk.CTkFrame(content, fg_color="transparent")
        body.grid(row=2, column=0, sticky="ew", padx=16, pady=(0, 8))
        body.grid_columnconfigure(0, weight=1)

        line_header = ctk.CTkFrame(body, fg_color="transparent")
        line_header.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        line_header.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(
            line_header,
            text="PRODUCT DETAILS",
            font=FONT_HEADING,
        ).grid(row=0, column=0, sticky="w")
        ctk.CTkLabel(
            line_header,
            text="Select an item in the blank row, then enter Qty and Price like an Excel sheet.",
            font=FONT_SMALL,
            text_color=COLORS["text_muted"],
        ).grid(row=1, column=0, sticky="w", pady=(2, 0))
        ctk.CTkButton(
            line_header,
            text="+ Add Product Row",
            width=150,
            fg_color=COLORS["success"],
            command=self._add_product_row,
        ).grid(row=0, column=1, rowspan=2, sticky="e")

        items_card = ctk.CTkFrame(body, fg_color=COLORS["bg"], corner_radius=10)
        items_card.grid(row=1, column=0, sticky="ew")
        items_card.grid_columnconfigure(0, weight=1)

        header = ctk.CTkFrame(
            items_card,
            fg_color="#F2AA2E",
            corner_radius=6,
        )
        header.grid(row=0, column=0, sticky="ew", padx=12, pady=(10, 4))
        header.grid_columnconfigure(0, weight=1)
        for col, text, width in [
            (0, "Product", 230),
            (1, "Stock", 90),
            (2, "Qty", 70),
            (3, "Price", 100),
            (4, "Tax %", 75),
            (5, "Disc %", 75),
            (6, "Amount", 120),
            (7, "", 45),
        ]:
            ctk.CTkLabel(
                header,
                text=text,
                font=FONT_HEADING,
                width=width,
                anchor="center" if col != 1 else "w",
                text_color="#292A33",
            ).grid(
                row=0, column=col, sticky="ew", padx=6, pady=9
            )

        items_viewport = ctk.CTkFrame(
            items_card,
            fg_color="transparent",
            height=250,
        )
        items_viewport.grid(row=1, column=0, sticky="ew", padx=12, pady=(0, 10))
        items_viewport.grid_propagate(False)
        items_viewport.grid_columnconfigure(0, weight=1)
        items_viewport.grid_rowconfigure(0, weight=1)

        self.items_container = ctk.CTkScrollableFrame(
            items_viewport,
            fg_color="transparent",
            scrollbar_button_color="#F2AA2E",
            scrollbar_button_hover_color="#D99116",
        )
        self.items_container.grid(
            row=0, column=0, sticky="nsew"
        )
        self.items_container.grid_columnconfigure(0, weight=1)

        # Excel-like totals rows at the end of the product table.
        totals = ctk.CTkFrame(
            items_card,
            fg_color=COLORS["card"],
            corner_radius=6,
            border_width=1,
            border_color=COLORS["border"],
        )
        totals.grid(row=2, column=0, sticky="ew", padx=12, pady=(0, 12))
        totals.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            totals, text="BILL SUMMARY", font=FONT_HEADING, anchor="w"
        ).grid(row=0, column=0, sticky="ew", padx=16, pady=(12, 8))
        self.subtotal_label = ctk.CTkLabel(
            totals, text="Subtotal: ₹0.00", font=FONT_BODY, anchor="e"
        )
        self.subtotal_label.grid(row=1, column=0, sticky="ew", padx=16, pady=3)
        self.discount_label = ctk.CTkLabel(
            totals, text="Discount: ₹0.00", font=FONT_BODY, anchor="e"
        )
        self.discount_label.grid(row=2, column=0, sticky="ew", padx=16, pady=3)
        self.tax_label = ctk.CTkLabel(
            totals, text="Tax: ₹0.00", font=FONT_BODY, anchor="e"
        )
        self.tax_label.grid(row=3, column=0, sticky="ew", padx=16, pady=3)

        self.total_label = ctk.CTkLabel(
            totals,
            text="Total Amount: ₹0.00",
            font=("Helvetica", 20, "bold"),
            text_color="#292A33",
            fg_color="#F2AA2E",
            corner_radius=4,
            anchor="e",
        )
        self.total_label.grid(row=4, column=0, sticky="ew", padx=16, pady=(8, 10), ipady=7)

        cash_row = ctk.CTkFrame(totals, fg_color="transparent")
        cash_row.grid(row=5, column=0, sticky="e", padx=16, pady=4)
        ctk.CTkLabel(
            cash_row, text="Cash Received", font=FONT_BODY, width=120, anchor="e"
        ).pack(side="left", padx=(0, 8))
        cash_entry = ctk.CTkEntry(
            cash_row, textvariable=self.cash_received_var, width=130,
            placeholder_text="0.00",
        )
        cash_entry.pack(side="left")
        cash_entry.bind("<KeyRelease>", lambda event: self._update_totals())
        self.balance_label = ctk.CTkLabel(
            totals, text="Balance Return: ₹0.00", font=FONT_HEADING, anchor="e"
        )
        self.balance_label.grid(row=6, column=0, sticky="ew", padx=16, pady=(4, 12))

        btn_row = ctk.CTkFrame(content, fg_color="transparent")
        btn_row.grid(row=3, column=0, sticky="e", padx=16, pady=(4, 16))
        ctk.CTkButton(
            btn_row,
            text="Generate Invoice",
            width=140,
            fg_color="#292A33",
            hover_color="#3F404A",
            command=self._print_bill,
        ).pack(side="left", padx=6)
        ctk.CTkButton(
            btn_row,
            text="Save Bill",
            width=140,
            fg_color=COLORS["success"],
            command=self._save_bill,
        ).pack(side="left", padx=6)
        ctk.CTkButton(
            btn_row,
            text="Download PDF",
            width=140,
            fg_color="#F2AA2E",
            hover_color="#D99116",
            text_color="#292A33",
            command=self._download_pdf,
        ).pack(side="left", padx=6)
        ctk.CTkButton(
            btn_row,
            text="Download Excel",
            width=140,
            fg_color="#F2AA2E",
            hover_color="#D99116",
            text_color="#292A33",
            command=self._download_excel,
        ).pack(side="left", padx=6)
        ctk.CTkButton(
            btn_row,
            text="Print Bill",
            width=140,
            fg_color="#292A33",
            hover_color="#3F404A",
            command=self._print_bill,
        ).pack(side="left", padx=6)

        self._render_item_table()

    def _add_manual_item(self, product, quantity, unit, unit_price):
        product_var = ctk.StringVar(value=product["name"])
        qty_var = ctk.StringVar(value=f"{quantity:g}")
        price_var = ctk.StringVar(value=f"{unit_price:.2f}")
        unit_var = ctk.StringVar(value=unit.strip() or product.get("unit", "pcs"))
        tax_var = ctk.StringVar(value=f"{float(db.get_settings().get('default_tax', 0)):g}")
        discount_var = ctk.StringVar(value="0")
        item = {
            "product_id": product["id"],
            "product_name": product["name"],
            "product_var": product_var,
            "unit": unit.strip() or product.get("unit", "pcs"),
            "quantity": quantity,
            "unit_price": unit_price,
            "qty_var": qty_var,
            "price_var": price_var,
            "unit_var": unit_var,
            "tax_var": tax_var,
            "discount_var": discount_var,
            "tax_rate": 0.0,
            "discount_rate": 0.0,
            "tax_amount": 0.0,
            "discount_amount": 0.0,
            "amount": quantity * unit_price,
            "max_stock": float(product["stock_quantity"]),
        }
        self.cart.append(item)
        self._render_item_table()

    def _add_product_row(self):
        products = [
            p for p in db.get_products()
            if float(p.get("stock_quantity", 0)) > 0
        ]
        if not products:
            messagebox.showwarning(
                "No products available",
                "Please add a product with available stock in the Inventory page first.",
            )
            return
        dialog = ctk.CTkToplevel(self)
        dialog.title("Add Product to Bill")
        dialog.geometry("500x520")
        dialog.resizable(False, False)
        dialog.transient(self.winfo_toplevel())
        dialog.grab_set()

        form = ctk.CTkFrame(dialog, fg_color=COLORS["card"], corner_radius=12)
        form.pack(fill="both", expand=True, padx=18, pady=18)
        form.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            form, text="Add Product to Bill", font=("Helvetica", 20, "bold")
        ).grid(row=0, column=0, sticky="w", padx=20, pady=(20, 4))
        ctk.CTkLabel(
            form,
            text="Choose a product and enter its billing values.",
            font=FONT_SMALL,
            text_color=COLORS["text_muted"],
        ).grid(row=1, column=0, sticky="w", padx=20, pady=(0, 14))

        product_map = {p["name"]: p for p in products}
        product_var = ctk.StringVar(value="")
        qty_var = ctk.StringVar(value="1")
        price_var = ctk.StringVar(value="")
        tax_var = ctk.StringVar(value="0")
        discount_var = ctk.StringVar(value="0")
        stock_var = ctk.StringVar(value="Select a product to see available stock.")

        def field(row, label, widget):
            ctk.CTkLabel(form, text=label, font=FONT_BODY).grid(
                row=row, column=0, sticky="w", padx=20, pady=(8, 3)
            )
            widget.grid(row=row + 1, column=0, sticky="ew", padx=20)

        product_combo = ctk.CTkComboBox(
            form,
            values=list(product_map.keys()),
            variable=product_var,
        )
        product_combo.set("Select product...")
        field(2, "Product *", product_combo)
        ctk.CTkLabel(
            form, textvariable=stock_var, font=FONT_SMALL,
            text_color=COLORS["success"],
        ).grid(row=4, column=0, sticky="w", padx=20, pady=(4, 2))

        field(5, "Quantity *", ctk.CTkEntry(form, textvariable=qty_var))
        field(7, "Price *", ctk.CTkEntry(form, textvariable=price_var))

        percentages = ctk.CTkFrame(form, fg_color="transparent")
        percentages.grid(row=9, column=0, sticky="ew", padx=20, pady=(8, 0))
        percentages.grid_columnconfigure(0, weight=1)
        percentages.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(percentages, text="Tax %", font=FONT_BODY).grid(
            row=0, column=0, sticky="w", padx=(0, 8)
        )
        ctk.CTkLabel(percentages, text="Discount %", font=FONT_BODY).grid(
            row=0, column=1, sticky="w", padx=(8, 0)
        )
        ctk.CTkEntry(percentages, textvariable=tax_var).grid(
            row=1, column=0, sticky="ew", padx=(0, 8)
        )
        ctk.CTkEntry(percentages, textvariable=discount_var).grid(
            row=1, column=1, sticky="ew", padx=(8, 0)
        )

        error_label = ctk.CTkLabel(
            form, text="", font=FONT_SMALL, text_color=COLORS["danger"]
        )
        error_label.grid(row=10, column=0, sticky="w", padx=20, pady=(10, 0))

        def select_product(selected_name):
            product = product_map.get(selected_name)
            if not product:
                return
            price_var.set(f"{float(product['selling_price']):.2f}")
            stock_var.set(
                f"Available: {float(product['stock_quantity']):g} "
                f"{product.get('unit', 'kg')}"
            )

        product_combo.configure(command=select_product)

        def add_to_bill():
            product = product_map.get(product_var.get())
            if not product:
                error_label.configure(text="Please select a product.")
                return
            try:
                quantity = float(qty_var.get())
                unit_price = float(
                    price_var.get() or product.get("selling_price", 0)
                )
                tax_rate = float(tax_var.get() or 0)
                discount_rate = float(discount_var.get() or 0)
                if quantity <= 0 or unit_price < 0:
                    raise ValueError
            except ValueError:
                error_label.configure(text="Enter valid quantity, price, tax and discount.")
                return
            already_entered = sum(
                item["quantity"] for item in self.cart
                if item["product_id"] == product["id"]
            )
            if quantity + already_entered > float(product["stock_quantity"]):
                error_label.configure(
                    text=f"Only {float(product['stock_quantity']):g} "
                    f"{product.get('unit', 'kg')} is available."
                )
                return
            self._add_manual_item(
                product, quantity, product.get("unit", "kg"), unit_price
            )
            item = self.cart[-1]
            item["tax_var"].set(f"{tax_rate:g}")
            item["discount_var"].set(f"{discount_rate:g}")
            self._sync_line_item(len(self.cart) - 1)
            dialog.destroy()

        ctk.CTkButton(
            form,
            text="Add to Bill",
            height=42,
            fg_color=COLORS["success"],
            command=add_to_bill,
        ).grid(row=11, column=0, sticky="ew", padx=20, pady=(12, 20))

    def _change_line_product(self, idx, selected_name):
        if idx >= len(self.cart):
            return
        product = next(
            (
                p for p in db.get_products(selected_name)
                if p["name"].lower() == selected_name.lower()
            ),
            None,
        )
        if not product or float(product.get("stock_quantity", 0)) <= 0:
            messagebox.showwarning(
                "Product unavailable",
                f"{selected_name} does not have available stock.",
            )
            return
        item = self.cart[idx]
        item["product_id"] = product["id"]
        item["product_name"] = product["name"]
        item["product_var"].set(product["name"])
        item["unit"] = product.get("unit", "kg")
        item["unit_var"].set(item["unit"])
        item["max_stock"] = float(product["stock_quantity"])
        item["unit_price"] = float(product.get("selling_price", 0))
        item["price_var"].set(f"{item['unit_price']:.2f}")
        if item["quantity"] > item["max_stock"]:
            item["quantity"] = item["max_stock"]
            item["qty_var"].set(f"{item['quantity']:g}")
        self._sync_line_item(idx)
        self._render_item_table()

    def _render_item_table(self):
        for widget in self.items_container.winfo_children():
            widget.destroy()

        if not self.cart:
            ctk.CTkLabel(
                self.items_container,
                text="Click “+ Add Product Row” to choose and add a product.",
                font=FONT_SMALL,
                text_color=COLORS["text_muted"],
            ).pack(pady=18)

        for idx, item in enumerate(self.cart):
            row = ctk.CTkFrame(
                self.items_container,
                fg_color=COLORS["card"],
                corner_radius=4,
                border_width=1,
                border_color=COLORS["border"],
            )
            row.pack(fill="x", pady=2)
            row.grid_columnconfigure(0, weight=0)
            row.grid_columnconfigure(0, weight=1)
            row.grid_columnconfigure(2, weight=0)
            row.grid_columnconfigure(3, weight=0)
            row.grid_columnconfigure(4, weight=0)
            row.grid_columnconfigure(6, weight=1)

            availability_color = (
                COLORS["warning"]
                if item["max_stock"] <= 5
                else COLORS["text_muted"]
            )
            product_cell = ctk.CTkFrame(row, fg_color="transparent", width=230)
            product_cell.grid(row=0, column=0, sticky="ew", padx=6, pady=4)
            available_names = [
                p["name"] for p in db.get_products()
                if float(p.get("stock_quantity", 0)) > 0
            ]
            ctk.CTkComboBox(
                product_cell,
                values=available_names,
                variable=item["product_var"],
                command=lambda value, i=idx: self._change_line_product(i, value),
            ).pack(fill="x")
            ctk.CTkLabel(
                row,
                text=f"{item['max_stock']:g} {item['unit']}",
                font=FONT_SMALL,
                width=90,
                text_color=availability_color,
            ).grid(row=0, column=1, padx=6, pady=4)
            qty_entry = ctk.CTkEntry(row, textvariable=item["qty_var"], width=70)
            qty_entry.grid(row=0, column=2, padx=(0, 6), pady=4)
            qty_entry.bind("<KeyRelease>", lambda e, i=idx: self._sync_line_item(i))
            qty_entry.bind("<FocusOut>", lambda e, i=idx: self._sync_line_item(i, True))
            price_entry = ctk.CTkEntry(row, textvariable=item["price_var"], width=100)
            price_entry.grid(row=0, column=3, padx=(0, 6), pady=4)
            price_entry.bind("<KeyRelease>", lambda e, i=idx: self._sync_line_item(i))
            tax_entry = ctk.CTkEntry(row, textvariable=item["tax_var"], width=75)
            tax_entry.grid(row=0, column=4, padx=(0, 6), pady=4)
            tax_entry.bind("<KeyRelease>", lambda e, i=idx: self._sync_line_item(i))
            discount_entry = ctk.CTkEntry(
                row, textvariable=item["discount_var"], width=75
            )
            discount_entry.grid(row=0, column=5, padx=(0, 6), pady=4)
            discount_entry.bind("<KeyRelease>", lambda e, i=idx: self._sync_line_item(i))
            amount_label = ctk.CTkLabel(
                row,
                text=format_currency(item["amount"]),
                font=FONT_MONO,
                width=120,
                anchor="w",
            )
            amount_label.grid(row=0, column=6, padx=(0, 6), pady=4)
            item["amount_label"] = amount_label
            ctk.CTkButton(
                row,
                text="✕",
                width=40,
                fg_color=COLORS["danger"],
                command=lambda i=idx: self._remove_item(i),
            ).grid(row=0, column=7, sticky="e", padx=(6, 0), pady=4)

        self._update_totals()

    def _sync_line_item(self, idx, show_warning=False):
        if idx >= len(self.cart):
            return
        item = self.cart[idx]
        try:
            quantity = float(item["qty_var"].get())
        except Exception:
            quantity = 0
        try:
            unit_price = float(item["price_var"].get())
        except Exception:
            unit_price = 0
        try:
            tax_rate = float(item["tax_var"].get())
        except Exception:
            tax_rate = 0
        try:
            discount_rate = float(item["discount_var"].get())
        except Exception:
            discount_rate = 0
        quantity = max(quantity, 0)
        if quantity > item["max_stock"]:
            quantity = item["max_stock"]
            item["qty_var"].set(f"{quantity:g}")
            if show_warning:
                messagebox.showwarning(
                    "Stock limit reached",
                    f"Only {quantity:g} {item['unit']} of {item['product_name']} is available.",
                )
        unit_price = max(unit_price, 0)
        tax_rate = max(tax_rate, 0)
        discount_rate = min(max(discount_rate, 0), 100)
        line_subtotal = quantity * unit_price
        discount_amount = line_subtotal * (discount_rate / 100)
        taxable_amount = line_subtotal - discount_amount
        tax_amount = taxable_amount * (tax_rate / 100)
        amount = taxable_amount + tax_amount
        item["quantity"] = quantity
        item["unit_price"] = unit_price
        item["tax_rate"] = tax_rate
        item["discount_rate"] = discount_rate
        item["discount_amount"] = discount_amount
        item["tax_amount"] = tax_amount
        item["amount"] = amount
        item["subtotal"] = line_subtotal
        if item.get("amount_label"):
            item["amount_label"].configure(text=format_currency(amount))
        self._update_totals()

    def _sync_unit(self, idx, value):
        if idx < len(self.cart):
            self.cart[idx]["unit"] = value

    def _remove_item(self, idx):
        if idx < len(self.cart):
            del self.cart[idx]
            self._render_item_table()

    def _update_totals(self):
        subtotal, _, tax_amount, discount, total = self._calc_totals()
        try:
            cash_received = float(self.cash_received_var.get() or 0)
        except Exception:
            cash_received = 0
        balance = max(cash_received - total, 0)

        self.subtotal_label.configure(text=f"Subtotal: {format_currency(subtotal)}")
        self.tax_label.configure(text=f"Tax: {format_currency(tax_amount)}")
        self.discount_label.configure(text=f"Discount: {format_currency(discount)}")
        self.total_label.configure(text=f"Total Amount: {format_currency(total)}")
        self.balance_label.configure(text=f"Balance Return: {format_currency(balance)}")

    def _apply_customer(self, event=None):
        selected = self.customer_combo.get()
        if selected == "Walk-in Customer":
            self.customer_name.set("")
            self.customer_phone.set("")
            self.customer_email.set("")
            self.customer_gst.set("")
            self.customer_address_box.delete("1.0", "end")
            self.tax_rate_var.set("0")
            return
        customer = db.get_customers(search=selected)
        if customer:
            customer = customer[0]
            self.customer_name.set(customer["name"])
            self.customer_phone.set(customer.get("phone", ""))
            self.customer_email.set(customer.get("email", ""))
            self.customer_gst.set(customer.get("gst", ""))
            self.customer_address_box.delete("1.0", "end")
            self.customer_address_box.insert("1.0", customer.get("address", ""))
            self.tax_rate_var.set("18" if customer.get("gst") else "0")

    def _calc_totals(self):
        subtotal = sum(
            item["quantity"] * item["unit_price"] for item in self.cart
        )
        discount = sum(item.get("discount_amount", 0) for item in self.cart)
        tax_amount = sum(item.get("tax_amount", 0) for item in self.cart)
        total = sum(item.get("amount", 0) for item in self.cart)
        return subtotal, 0, tax_amount, discount, total

    def _save_bill(self):
        if not self.cart:
            messagebox.showwarning("Warning", "Please add at least one product before saving the bill.")
            return

        customer_name = self.customer_name.get().strip()
        customer_gst = self.customer_gst.get().strip()
        customer_phone = self.customer_phone.get().strip()
        customer_email = self.customer_email.get().strip()
        customer_address = self.customer_address_box.get("1.0", "end").strip()
        subtotal, tax_rate, tax_amount, discount, total = self._calc_totals()
        try:
            cash_received = float(self.cash_received_var.get() or 0)
        except ValueError:
            cash_received = 0

        if any(item["quantity"] <= 0 for item in self.cart):
            messagebox.showwarning("Invalid quantity", "Every line item must have a quantity above zero.")
            return

        # Refresh availability immediately before committing the invoice.
        for item in self.cart:
            product = db.get_product(item["product_id"])
            if not product:
                messagebox.showerror("Error", f"Product {item['product_name']} no longer exists.")
                return
            if item["quantity"] > product["stock_quantity"]:
                messagebox.showerror(
                    "Insufficient stock",
                    f"Only {product['stock_quantity']} units available for {item['product_name']}.",
                )
                return

        customer_id = self._customers.get(self.customer_combo.get())
        if self.save_customer_var.get() and customer_name and not customer_id:
            customer_id = db.add_customer({
                "name": customer_name,
                "phone": customer_phone,
                "email": customer_email,
                "gst": customer_gst,
                "address": customer_address,
            })

        sale_data = {
            "invoice_number": self.invoice_number.get().strip() or db.generate_invoice_number(),
            "customer_id": customer_id,
            "customer_name": customer_name or "Walk-in Customer",
            "customer_phone": customer_phone,
            "customer_email": customer_email,
            "customer_gst": customer_gst,
            "customer_address": customer_address,
            "subtotal": subtotal,
            "tax_rate": tax_rate,
            "tax_amount": tax_amount,
            "discount": discount,
            "total": total,
            "payment_method": self.payment_var.get(),
            "cash_received": cash_received,
            "balance_return": max(cash_received - total, 0),
        }
        try:
            sale_id = db.create_sale(sale_data, self.cart)
        except ValueError as exc:
            messagebox.showerror("Unable to save bill", str(exc))
            return
        except Exception as exc:
            messagebox.showerror("Unable to save bill", f"The invoice could not be saved.\n{exc}")
            return

        export_dir = os.path.join(os.getcwd(), "exports")
        os.makedirs(export_dir, exist_ok=True)
        file_name = f"bill_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        file_path = os.path.join(export_dir, file_name)
        with open(file_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Invoice", sale_data["invoice_number"]])
            writer.writerow(["Customer", customer_name or "Walk-in Customer"])
            writer.writerow(["Phone", customer_phone])
            writer.writerow(["Email", customer_email])
            writer.writerow(["GST", customer_gst])
            writer.writerow(["Address", customer_address])
            writer.writerow([])
            writer.writerow([
                "Product", "Stock", "Qty", "Price", "Tax %", "Disc %", "Amount"
            ])
            for idx, item in enumerate(self.cart, start=1):
                writer.writerow([
                    item["product_name"],
                    item["max_stock"],
                    item["quantity"],
                    item["unit_price"],
                    item.get("tax_rate", 0),
                    item.get("discount_rate", 0),
                    item.get("amount", 0),
                ])
            writer.writerow([])
            writer.writerow(["Subtotal", format_currency(subtotal)])
            writer.writerow(["Discount", format_currency(discount)])
            writer.writerow(["Tax", format_currency(tax_amount)])
            writer.writerow(["Grand Total", format_currency(total)])
            writer.writerow(["Cash Received", format_currency(cash_received)])
            writer.writerow(["Balance Return", format_currency(max(cash_received - total, 0))])

        messagebox.showinfo(
            "Saved",
            f"Bill saved successfully.\nInvoice: {sale_data['invoice_number']}\nCSV file: {file_path}",
        )
        self.cart.clear()
        self.refresh()
        if hasattr(self, "on_sale_complete"):
            self.on_sale_complete(sale_id)

    def _download_pdf(self):
        if not self.cart:
            messagebox.showwarning("Warning", "Please add at least one product before exporting PDF.")
            return

        if not all([Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle, rl_colors, A4]):
            messagebox.showerror(
                "Missing dependency",
                "PDF export requires the 'reportlab' package. Please install it and try again.",
            )
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            initialfile=f"invoice_{self.invoice_number.get().replace('/', '_')}.pdf",
        )
        if not file_path:
            return
        self._build_pdf_file(file_path)
        messagebox.showinfo("Exported", f"PDF saved to:\n{file_path}")

    def _build_pdf_file(self, file_path):
        subtotal, _, tax_amount, discount, total = self._calc_totals()
        doc = SimpleDocTemplate(
            file_path,
            pagesize=A4,
            rightMargin=18 * mm,
            leftMargin=18 * mm,
            topMargin=16 * mm,
            bottomMargin=16 * mm,
        )
        styles = getSampleStyleSheet()
        story = []
        story.append(Paragraph("INVOICE", styles["Title"]))
        story.append(Paragraph(f"Invoice: {self.invoice_number.get()}", styles["Heading2"]))
        story.append(Paragraph(f"Date: {self.invoice_date.get()}", styles["BodyText"]))
        story.append(Spacer(1, 8 * mm))

        customer_lines = [
            f"Customer: {self.customer_name.get() or 'Walk-in Customer'}",
            f"Phone: {self.customer_phone.get()}",
            f"Email: {self.customer_email.get()}",
            f"GST: {self.customer_gst.get()}",
            f"Address: {self.customer_address_box.get('1.0', 'end').strip()}",
        ]
        story.append(Paragraph("Bill To", styles["Heading3"]))
        for line in customer_lines:
            story.append(Paragraph(line, styles["BodyText"]))
        story.append(Spacer(1, 6 * mm))

        table_data = [["Product", "Stock", "Qty", "Price", "Tax %", "Disc %", "Amount"]]
        for item in self.cart:
            table_data.append([
                item["product_name"],
                f"{item['max_stock']:g}",
                f"{item['quantity']:g}",
                f"INR {item['unit_price']:.2f}",
                f"{item.get('tax_rate', 0):g}",
                f"{item.get('discount_rate', 0):g}",
                f"INR {item.get('amount', 0):.2f}",
            ])
        table = Table(
            table_data,
            repeatRows=1,
            colWidths=[45*mm, 20*mm, 18*mm, 25*mm, 18*mm, 18*mm, 28*mm],
        )
        table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#2563EB")),
                ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.whitesmoke),
                ("GRID", (0, 0), (-1, -1), 0.5, rl_colors.grey),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (0, 1), (0, -1), "CENTER"),
                ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
            ])
        )
        story.append(table)
        story.append(Spacer(1, 8 * mm))

        totals = [
            ("Subtotal", f"INR {subtotal:,.2f}"),
            ("Discount", f"INR {discount:,.2f}"),
            ("Tax", f"INR {tax_amount:,.2f}"),
            ("Grand Total", f"INR {total:,.2f}"),
        ]
        for label, value in totals:
            story.append(
                Paragraph(
                    f"<b>{label}:</b> {value}",
                    styles["BodyText"],
                )
            )
        story.append(Spacer(1, 10 * mm))
        story.append(Paragraph(
            f"Payment method: {self.payment_var.get()}", styles["BodyText"]
        ))
        try:
            cash_received = float(self.cash_received_var.get() or 0)
        except ValueError:
            cash_received = 0
        story.append(Paragraph(
            f"Cash received: INR {cash_received:,.2f}", styles["BodyText"]
        ))
        story.append(Paragraph(
            f"Balance return: INR {max(cash_received - total, 0):,.2f}",
            styles["BodyText"],
        ))
        story.append(Spacer(1, 8 * mm))
        doc.build(story)

    def _download_excel(self):
        if not self.cart:
            messagebox.showwarning("Warning", "Please add at least one product before exporting Excel.")
            return
        if Workbook is None:
            messagebox.showerror(
                "Missing dependency",
                "Excel export requires the 'openpyxl' package. Please install it and try again.",
            )
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"invoice_{self.invoice_number.get().replace('/', '_')}.xlsx",
        )
        if not file_path:
            return

        subtotal, tax_rate, tax_amount, discount, total = self._calc_totals()
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice"
        ws.append(["INVOICE"])
        ws.append([f"Invoice: {self.invoice_number.get()} ", f"Date: {self.invoice_date.get()}"])
        ws.append([])
        ws.append(["Customer", self.customer_name.get() or "Walk-in Customer"])
        ws.append(["Phone", self.customer_phone.get()])
        ws.append(["Email", self.customer_email.get()])
        ws.append(["GST", self.customer_gst.get()])
        ws.append(["Address", self.customer_address_box.get("1.0", "end").strip()])
        ws.append([])
        ws.append(["Product", "Stock", "Qty", "Price", "Tax %", "Disc %", "Amount"])
        for item in self.cart:
            ws.append([
                item["product_name"],
                item["max_stock"],
                item["quantity"],
                item["unit_price"],
                item.get("tax_rate", 0),
                item.get("discount_rate", 0),
                item.get("amount", 0),
            ])
        ws.append([])
        ws.append(["Subtotal", format_currency(subtotal)])
        ws.append(["Discount", format_currency(discount)])
        ws.append(["Tax", format_currency(tax_amount)])
        ws.append(["Grand Total", format_currency(total)])
        try:
            cash_received = float(self.cash_received_var.get() or 0)
        except ValueError:
            cash_received = 0
        ws.append(["Cash Received", format_currency(cash_received)])
        ws.append(["Balance Return", format_currency(max(cash_received - total, 0))])
        for column, width in {
            "A": 28, "B": 12, "C": 12, "D": 16,
            "E": 12, "F": 12, "G": 18,
        }.items():
            ws.column_dimensions[column].width = width
        ws.freeze_panes = "A11"
        header_fill = PatternFill("solid", fgColor="2563EB")
        white_bold = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="CBD5E1")
        item_header_row = 10
        for cell in ws[item_header_row]:
            cell.fill = header_fill
            cell.font = white_bold
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        for row in ws.iter_rows(
            min_row=item_header_row + 1,
            max_row=item_header_row + len(self.cart),
            min_col=1,
            max_col=7,
        ):
            for cell in row:
                cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        ws["A1"].font = Font(size=18, bold=True, color="166534")
        ws.merge_cells("A1:G1")
        ws["A1"].alignment = Alignment(horizontal="center")
        wb.save(file_path)
        messagebox.showinfo("Exported", f"Excel saved to:\n{file_path}")

    def _print_bill(self):
        if not self.cart:
            messagebox.showwarning("Warning", "Please add at least one product before printing.")
            return

        subtotal, tax_rate, tax_amount, discount, total = self._calc_totals()
        preview = ctk.CTkToplevel(self)
        preview.title("Bill Preview")
        preview.geometry("600x700")
        preview.resizable(False, False)
        preview.transient(self.winfo_toplevel())
        preview.grab_set()

        content = ctk.CTkScrollableFrame(preview, fg_color=COLORS["card"], width=560, height=660)
        content.pack(fill="both", expand=True, padx=16, pady=16)

        ctk.CTkLabel(
            content, text="INVOICE", font=("Helvetica", 22, "bold")
        ).pack(anchor="w", pady=(8, 4))
        ctk.CTkLabel(content, text=f"Invoice: {self.invoice_number.get()}", font=FONT_HEADING).pack(anchor="w")
        ctk.CTkLabel(content, text=f"Date: {self.invoice_date.get()}", font=FONT_SMALL).pack(anchor="w", pady=(0, 10))

        ctk.CTkLabel(content, text="Bill To:", font=FONT_HEADING).pack(anchor="w")
        ctk.CTkLabel(
            content,
            text=self.customer_name.get() or "Walk-in Customer",
            font=FONT_BODY,
        ).pack(anchor="w")
        ctk.CTkLabel(content, text=self.customer_phone.get(), font=FONT_SMALL).pack(anchor="w")
        ctk.CTkLabel(content, text=self.customer_email.get(), font=FONT_SMALL).pack(anchor="w")
        ctk.CTkLabel(content, text=self.customer_gst.get() or "", font=FONT_SMALL).pack(anchor="w")
        ctk.CTkLabel(content, text=self.customer_address_box.get("1.0", "end").strip(), font=FONT_SMALL).pack(anchor="w", pady=(0, 10))

        header = ctk.CTkFrame(content, fg_color="transparent")
        header.pack(fill="x", pady=(0, 4))
        for text, width in [
            ("Product", 150), ("Stock", 55), ("Qty", 45), ("Price", 70),
            ("Tax", 45), ("Disc", 45), ("Amount", 85)
        ]:
            ctk.CTkLabel(header, text=text, font=FONT_HEADING, width=width, anchor="w").pack(side="left")

        for item in self.cart:
            row = ctk.CTkFrame(content, fg_color="transparent")
            row.pack(fill="x", pady=2)
            for text, width in [
                (item["product_name"], 150),
                (f"{item['max_stock']:g}", 55),
                (f"{item['quantity']:g}", 45),
                (format_currency(item["unit_price"]), 70),
                (f"{item.get('tax_rate', 0):g}%", 45),
                (f"{item.get('discount_rate', 0):g}%", 45),
                (format_currency(item.get("amount", 0)), 85),
            ]:
                ctk.CTkLabel(
                    row, text=text, font=FONT_MONO, width=width, anchor="w"
                ).pack(side="left")

        ctk.CTkFrame(content, height=2, fg_color=COLORS["border"]).pack(fill="x", pady=12)
        for label, value in [
            ("Subtotal", format_currency(subtotal)),
            ("Discount", format_currency(discount)),
            ("Tax", format_currency(tax_amount)),
            ("Grand Total", format_currency(total)),
        ]:
            row = ctk.CTkFrame(content, fg_color="transparent")
            row.pack(fill="x", pady=1)
            ctk.CTkLabel(row, text=label, font=FONT_BODY, anchor="e", width=180).pack(side="left")
            ctk.CTkLabel(row, text=value, font=FONT_BODY).pack(side="right")

        ctk.CTkLabel(content, text="Payment Method: " + self.payment_var.get(), font=FONT_SMALL).pack(anchor="w", pady=(12, 6))
        try:
            cash_received = float(self.cash_received_var.get() or 0)
        except ValueError:
            cash_received = 0
        ctk.CTkLabel(
            content, text=f"Cash Received: {format_currency(cash_received)}", font=FONT_SMALL
        ).pack(anchor="w")
        ctk.CTkLabel(
            content,
            text=f"Balance Return: {format_currency(max(cash_received - total, 0))}",
            font=FONT_HEADING,
        ).pack(anchor="w", pady=(2, 8))
        def send_to_printer():
            if not all([Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle, rl_colors, A4]):
                messagebox.showerror("Missing dependency", "Printing requires the reportlab package.")
                return
            temp_path = os.path.join(
                tempfile.gettempdir(),
                f"{self.invoice_number.get().replace('/', '_')}.pdf",
            )
            self._build_pdf_file(temp_path)
            try:
                subprocess.run(["lp", temp_path], check=True, capture_output=True, text=True)
                messagebox.showinfo("Print sent", "The bill was sent to the default printer.")
            except (FileNotFoundError, subprocess.CalledProcessError):
                subprocess.run(["open", temp_path], check=False)
                messagebox.showinfo(
                    "Print preview opened",
                    "No default printer was available, so the printable PDF was opened instead.",
                )

        button_row = ctk.CTkFrame(preview, fg_color="transparent")
        button_row.pack(pady=(0, 12))
        ctk.CTkButton(
            button_row, text="Print", width=110, fg_color=COLORS["success"],
            command=send_to_printer,
        ).pack(side="left", padx=5)
        ctk.CTkButton(
            button_row, text="Close", width=100, command=preview.destroy
        ).pack(side="left", padx=5)
