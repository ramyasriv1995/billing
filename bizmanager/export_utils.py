"""In-memory invoice, report, and database exports for the Streamlit UI."""

from io import BytesIO

from bson import json_util
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def invoice_pdf(invoice, items):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4, rightMargin=18 * mm, leftMargin=18 * mm,
        topMargin=16 * mm, bottomMargin=16 * mm,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph("INVOICE", styles["Title"]),
        Paragraph(f"Invoice: {invoice['invoice_number']}", styles["Heading2"]),
        Paragraph(f"Date: {invoice['invoice_date']}", styles["BodyText"]),
        Spacer(1, 8 * mm),
        Paragraph("Bill To", styles["Heading3"]),
    ]
    for label, value in [
        ("Customer", invoice.get("customer_name") or "Walk-in Customer"),
        ("Phone", invoice.get("customer_phone", "")),
        ("Email", invoice.get("customer_email", "")),
        ("GST", invoice.get("customer_gst", "")),
        ("Address", invoice.get("customer_address", "")),
    ]:
        story.append(Paragraph(f"{label}: {value}", styles["BodyText"]))
    story.append(Spacer(1, 6 * mm))
    rows = [["Product", "Stock", "Qty", "Price", "Tax %", "Disc %", "Amount"]]
    rows.extend([
        [
            item["product_name"], f"{item.get('max_stock', 0):g}",
            f"{item['quantity']:g}", f"INR {item['unit_price']:.2f}",
            f"{item.get('tax_rate', 0):g}", f"{item.get('discount_rate', 0):g}",
            f"INR {item.get('amount', 0):.2f}",
        ]
        for item in items
    ])
    table = Table(
        rows, repeatRows=1,
        colWidths=[45 * mm, 20 * mm, 18 * mm, 25 * mm, 18 * mm, 18 * mm, 28 * mm],
    )
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
    ]))
    story.extend([table, Spacer(1, 8 * mm)])
    for label, value in [
        ("Subtotal", invoice["subtotal"]),
        ("Discount", invoice["discount"]),
        ("Tax", invoice["tax_amount"]),
        ("Grand Total", invoice["total"]),
        ("Cash Received", invoice.get("cash_received", 0)),
        ("Balance Return", invoice.get("balance_return", 0)),
    ]:
        story.append(Paragraph(f"<b>{label}:</b> INR {value:,.2f}", styles["BodyText"]))
    story.append(Paragraph(
        f"Payment method: {invoice.get('payment_method', 'Cash')}", styles["BodyText"]
    ))
    doc.build(story)
    return buffer.getvalue()


def invoice_excel(invoice, items):
    buffer = BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Invoice"
    sheet.append(["INVOICE"])
    sheet.append([
        f"Invoice: {invoice['invoice_number']}",
        f"Date: {invoice['invoice_date']}",
    ])
    sheet.append([])
    for label, key, fallback in [
        ("Customer", "customer_name", "Walk-in Customer"),
        ("Phone", "customer_phone", ""),
        ("Email", "customer_email", ""),
        ("GST", "customer_gst", ""),
        ("Address", "customer_address", ""),
    ]:
        sheet.append([label, invoice.get(key) or fallback])
    sheet.append([])
    sheet.append(["Product", "Stock", "Qty", "Price", "Tax %", "Disc %", "Amount"])
    for item in items:
        sheet.append([
            item["product_name"], item.get("max_stock", 0), item["quantity"],
            item["unit_price"], item.get("tax_rate", 0),
            item.get("discount_rate", 0), item.get("amount", 0),
        ])
    sheet.append([])
    for label, key in [
        ("Subtotal", "subtotal"), ("Discount", "discount"), ("Tax", "tax_amount"),
        ("Grand Total", "total"), ("Cash Received", "cash_received"),
        ("Balance Return", "balance_return"),
    ]:
        sheet.append([label, invoice.get(key, 0)])
    for column, width in {
        "A": 28, "B": 18, "C": 12, "D": 16, "E": 12, "F": 12, "G": 18,
    }.items():
        sheet.column_dimensions[column].width = width
    header_row = 10
    fill = PatternFill("solid", fgColor="2563EB")
    white_bold = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="CBD5E1")
    for cell in sheet[header_row]:
        cell.fill = fill
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    sheet["A1"].font = Font(size=18, bold=True, color="166534")
    sheet.merge_cells("A1:G1")
    sheet["A1"].alignment = Alignment(horizontal="center")
    workbook.save(buffer)
    return buffer.getvalue()


def report_rows(report):
    if report["report_type"] == "Sales Report":
        return ["Product", "Quantity", "Revenue", "Profit"], [
            [r["product_name"], r["quantity"], r["revenue"], r["profit"]]
            for r in report["product_sales"]
        ]
    if report["report_type"] == "Inventory Report":
        return ["Product", "Stock", "Unit", "Inventory Value"], [
            [p["name"], p["stock_quantity"], p.get("unit", ""),
             p["stock_quantity"] * p["cost_price"]]
            for p in report["current_stock"]
        ]
    return ["Method", "Collection"], list(report["collections"].items())


def report_excel(report):
    headers, rows = report_rows(report)
    buffer = BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Report"
    sheet.append([report["report_type"]])
    sheet.append([report["period_label"]])
    sheet.append([])
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    for cell in sheet[4]:
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.font = Font(color="FFFFFF", bold=True)
    workbook.save(buffer)
    return buffer.getvalue()


def report_pdf(report):
    headers, rows = report_rows(report)
    buffer = BytesIO()
    styles = getSampleStyleSheet()
    story = [
        Paragraph(report["report_type"], styles["Title"]),
        Paragraph(report["period_label"], styles["BodyText"]),
        Spacer(1, 16),
    ]
    table = Table([headers] + rows, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))
    story.append(table)
    SimpleDocTemplate(buffer, pagesize=A4).build(story)
    return buffer.getvalue()


def database_json(database):
    payload = {
        name: list(database[name].find())
        for name in database.list_collection_names()
    }
    return json_util.dumps(payload, indent=2).encode("utf-8")
